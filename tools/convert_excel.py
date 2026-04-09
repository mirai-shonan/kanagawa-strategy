#!/usr/bin/env python3
"""
選挙管理委員会 Excel → CSV 変換ツール

prefectures/<県名>/raw/ にあるExcelファイルを読み込み、
prefectures/<県名>/data/ にCSVを出力します。

使い方:
  python3 tools/convert_excel.py prefectures/kanagawa

raw/ ディレクトリの命名規則（ファイル名にこれらのキーワードを含める）:
  - sangi + senkyoku                → 参院選挙区（候補者別開票区別）→ candidate.csv
  - sangi + hirei + district        → 参院比例（党派別開票区別）   → senate.csv
  - shugi + hirei + district        → 衆院比例（党派別開票区別）   → house.csv ※選管公開なしの可能性あり

注意:
  - 衆院比例の「党派別開票区別」データは多くの自治体で公開されておらず、
    PDFのみの場合があります。その場合は手作業で house.csv を作成してください。

必要パッケージ:
  pip install openpyxl
"""

import argparse
import csv
import glob
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl が必要です。 pip install openpyxl でインストールしてください。", file=sys.stderr)
    sys.exit(1)


# ============================================================
# 共通ヘルパー
# ============================================================

# 政令指定都市（合計行はスキップし、区を子として採用）
DESIGNATED_CITIES = {
    '札幌市', '仙台市', 'さいたま市', '千葉市', '川崎市', '横浜市', '相模原市',
    '新潟市', '静岡市', '浜松市', '名古屋市', '京都市', '大阪市', '堺市',
    '神戸市', '岡山市', '広島市', '北九州市', '福岡市', '熊本市',
}

# 郡名パターン（神奈川県以外も含む一般的なパターン）
GUN_RE = re.compile(r'^[^\s]+郡')


def is_total_row(name):
    """合計行かどうかを判定"""
    s = str(name).strip()
    return s.endswith('計') or s in ('合計', '指定都市計', 'その他の市計', '町村計', '県計')


def is_gun(name):
    """郡名のみの行か（〜郡で終わる）"""
    s = str(name).strip()
    return bool(GUN_RE.match(s)) and not any(c in s for c in '町村市区')


def strip_gun_prefix(name):
    """町村名から郡名を除去する（例: 「足柄下郡箱根町」→「箱根町」）"""
    return GUN_RE.sub('', name)


def parse_number(val):
    """セル値を数値に変換する。文字列の場合はカンマ・空白を除去して変換。"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        try:
            return float(val.strip().replace(',', '').replace(' ', '').replace('\u3000', ''))
        except ValueError:
            return None
    return None


def find_col_by_keyword(ws, keyword, rows=(5, 7, 11)):
    """指定された行から、キーワードを含む列インデックスを探す"""
    for row in rows:
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val and keyword in str(val):
                return col
    return None


def find_total_col(ws, rows=(5, 7, 8, 11)):
    """合計列を探す"""
    for row in rows:
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                s = str(val).strip()
                if '得票数計' in s or '合計' in s or s == '計':
                    return col
    return None


# ============================================================
# 参院選挙区（候補者別開票区別）パーサー — g.xlsx形式
# ============================================================

def extract_area_name(col_a, col_b):
    """
    col_a と col_b から市区町村名を抽出する。
    返り値: 採用すべき地域名（政令指定都市の親行や郡の親行・合計行はNone）
    """
    a = str(col_a).strip() if col_a else ''
    b = str(col_b).strip() if col_b else ''

    # 合計行はスキップ
    if a and is_total_row(a):
        return None
    if b and is_total_row(b):
        return None

    # 政令指定都市の親行はスキップ（区が子として後続）
    if a and a in DESIGNATED_CITIES:
        return None

    # 郡名のみの親行はスキップ（町村が子として後続）
    if a and is_gun(a):
        return None

    # col_b に値がある場合（政令指定都市の区 or 郡の町村）
    if b:
        return strip_gun_prefix(b)

    # col_a に値があり、政令指定都市・郡・合計でない → 一般市
    if a:
        return strip_gun_prefix(a)

    return None


def parse_senate_district(filepath):
    """
    参院選挙区の候補者別開票区別Excel (g.xlsx形式) をパースする。
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    team_col = find_col_by_keyword(ws, 'チームみらい', rows=(5, 7))
    if team_col is None:
        return None
    total_col = find_total_col(ws, rows=(5, 7, 8))

    results = []
    for row in range(9, ws.max_row + 1):
        area = extract_area_name(
            ws.cell(row=row, column=1).value,
            ws.cell(row=row, column=2).value,
        )
        if not area:
            continue

        team_votes = parse_number(ws.cell(row=row, column=team_col).value)
        if team_votes is None:
            continue
        team_votes = round(team_votes)

        total = parse_number(ws.cell(row=row, column=total_col).value) if total_col else None
        rate = round(team_votes / total * 100, 1) if total and total > 0 else 0

        results.append({
            '地域': area,
            '得票': team_votes,
            '率': rate,
        })

    return results


# ============================================================
# 参院比例（党派別開票区別）パーサー — m.xlsx形式
# ============================================================

def parse_senate_proportional(filepath):
    """
    参院比例の党派別開票区別Excel (m.xlsx形式) をパースする。

    想定フォーマット:
    - Sheet: '開票区別'
    - Row 7: 党派名（各党3列ずつ: 得票総数, 政党等の得票総数, 名簿登載者を除く得票総数）
    - Row 12+: データ行
      - Col A: 親市区名（政令指定都市の合計行）
      - Col B: 市区町村名（フルネーム）
      - 各党の最初の列: 得票総数
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['開票区別'] if '開票区別' in wb.sheetnames else wb.active

    # チームみらいの列を探す（row 7）
    team_col = find_col_by_keyword(ws, 'チームみらい', rows=(7,))
    if team_col is None:
        return None

    # 全政党の「得票総数」列を集めて自治体別合計を算出する
    # row 7 で値があるセルが「政党の先頭列」
    party_cols = []
    for col in range(3, ws.max_column + 1):
        val = ws.cell(row=7, column=col).value
        if val and str(val).strip() and str(val).strip() not in ('党派名',):
            party_cols.append(col)

    results = []
    for row in range(12, ws.max_row + 1):
        area = extract_area_name(
            ws.cell(row=row, column=1).value,
            ws.cell(row=row, column=2).value,
        )
        if not area:
            continue

        team_votes = parse_number(ws.cell(row=row, column=team_col).value)
        if team_votes is None:
            continue
        team_votes = round(team_votes)

        # 全政党の得票総数を合計
        total = 0
        for pc in party_cols:
            v = parse_number(ws.cell(row=row, column=pc).value)
            if v is not None:
                total += v

        rate = round(team_votes / total * 100, 1) if total > 0 else 0

        results.append({
            '地域': area,
            '得票数': team_votes,
            '得票率': rate,
        })

    return results


# ============================================================
# CSV出力 + ファイル探索
# ============================================================

def write_csv(rows, filepath, fieldnames):
    os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, '') for k in fieldnames})
    print(f"  → {filepath} ({len(rows)} rows)", file=sys.stderr)


def find_excel_files(raw_dir, *required_keywords, exclude_keywords=()):
    """raw_dir 内で全ての必須キーワードを含み、除外キーワードを含まないExcelファイルを探す"""
    matches = []
    for path in sorted(glob.glob(os.path.join(raw_dir, '*.xlsx')) + glob.glob(os.path.join(raw_dir, '*.xls'))):
        name = os.path.basename(path).lower()
        if all(kw.lower() in name for kw in required_keywords):
            if not any(ex.lower() in name for ex in exclude_keywords):
                matches.append(path)
    return matches


def convert_prefecture(prefecture_dir):
    raw_dir = os.path.join(prefecture_dir, 'raw')
    data_dir = os.path.join(prefecture_dir, 'data')

    if not os.path.isdir(raw_dir):
        print(f"Error: raw/ ディレクトリが見つかりません: {raw_dir}", file=sys.stderr)
        sys.exit(1)

    print(f"Processing: {prefecture_dir}", file=sys.stderr)
    print(f"  raw/  → {raw_dir}", file=sys.stderr)
    print(f"  data/ → {data_dir}", file=sys.stderr)
    print()

    converted = 0

    # 参院選挙区（候補者別） → candidate.csv
    files = find_excel_files(raw_dir, 'sangi', 'senkyoku')
    for f in files:
        print(f"参院選挙区: {os.path.basename(f)}", file=sys.stderr)
        results = parse_senate_district(f)
        if results:
            write_csv(results, os.path.join(data_dir, 'candidate.csv'), ['地域', '得票', '率'])
            converted += 1
        else:
            print(f"  Warning: パース失敗（チームみらい列が見つかりません）", file=sys.stderr)

    # 参院比例（党派別開票区別） → senate.csv
    files = find_excel_files(raw_dir, 'sangi', 'hirei', 'district')
    for f in files:
        print(f"参院比例（党派別開票区別）: {os.path.basename(f)}", file=sys.stderr)
        results = parse_senate_proportional(f)
        if results:
            write_csv(results, os.path.join(data_dir, 'senate.csv'), ['地域', '得票数', '得票率'])
            converted += 1
        else:
            print(f"  Warning: パース失敗", file=sys.stderr)

    # 衆院比例（党派別開票区別） → house.csv
    files = find_excel_files(raw_dir, 'shugi', 'hirei', 'district')
    for f in files:
        print(f"衆院比例（党派別開票区別）: {os.path.basename(f)}", file=sys.stderr)
        # 構造は参院比例と同様と想定（要検証）
        results = parse_senate_proportional(f)
        if results:
            write_csv(results, os.path.join(data_dir, 'house.csv'),
                      ['地域', '得票数', '得票率'])
            converted += 1
        else:
            print(f"  Warning: パース失敗", file=sys.stderr)

    if not find_excel_files(raw_dir, 'shugi', 'hirei', 'district'):
        print("Note: 衆院比例（党派別開票区別）のExcelが見つかりません。", file=sys.stderr)
        print("      多くの自治体ではPDFのみ公開のため、house.csv は手作業で作成してください。", file=sys.stderr)

    if converted == 0:
        print("\nWarning: 変換可能なExcelファイルが見つかりませんでした。", file=sys.stderr)
        sys.exit(1)

    print(f"\nDone! {converted} file(s) converted.", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(
        description='選挙管理委員会 Excel → CSV 変換ツール',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用例:
  python3 tools/convert_excel.py prefectures/kanagawa

ディレクトリ構造:
  prefectures/kanagawa/
  ├── raw/                                       ← 選管Excelをここに配置
  │   ├── R7.7_sangi_senkyoku.xlsx               → candidate.csv
  │   ├── R7.7_sangi_hirei_district.xlsx         → senate.csv
  │   └── R8.2_shugi_hirei_district.xlsx         → house.csv（公開あれば）
  └── data/                                      ← CSV自動生成
      ├── candidate.csv
      ├── senate.csv
      └── house.csv（手動 or 自動）

ファイル命名規則（ファイル名にキーワードを含める）:
  - sangi + senkyoku           → 参院選挙区候補者別      → candidate.csv
  - sangi + hirei + district   → 参院比例党派別開票区別  → senate.csv
  - shugi + hirei + district   → 衆院比例党派別開票区別  → house.csv
        """,
    )
    parser.add_argument(
        'prefecture_dir',
        help='都道府県ディレクトリ (例: prefectures/kanagawa)',
    )

    args = parser.parse_args()
    convert_prefecture(args.prefecture_dir)


if __name__ == '__main__':
    main()
